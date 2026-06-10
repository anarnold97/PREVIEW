#!/usr/bin/env python3
"""
DITA Compliance Report Generator — MTV / Migration Toolkit for Virtualization

Runs all standard DITA migration checks against documentation/ in the
forklift-documentation repository and writes results to an Excel workbook
with one sheet per check.

Requirements:
    pip install openpyxl

Usage (run from the root of the forklift-documentation repo):
    python3 dita-report-mtv.py
    python3 dita-report-mtv.py --vale ~/dita.ini --out mtv-dita-report.xlsx
"""

import argparse
import re
import subprocess
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Missing dependency — run: pip install openpyxl")


FIND_CMD = (
    "find documentation/ -type f -name '*.adoc'"
    " -not -name 'common-attributes.adoc'"
    " -not -name 'master.adoc'"
)
GREP_TARGETS = "documentation/"


# ---------------------------------------------------------------------------
# Shell helpers
# ---------------------------------------------------------------------------

def run(cmd: str) -> str:
    result = subprocess.run(
        cmd, shell=True, capture_output=True, text=True, errors="replace"
    )
    return result.stdout


def run_vale(vale_config: str, filter_expr: str = "") -> list[dict]:
    filter_flag = f"--filter '{filter_expr}'" if filter_expr else ""
    cmd = (
        f"{FIND_CMD} | xargs -n 1 -P 8 vale"
        f" --config {vale_config}"
        f" {filter_flag}"
        f" --output line"
    )
    rows = []
    for line in run(cmd).splitlines():
        parsed = _parse_vale_line(line)
        if parsed:
            rows.append(parsed)
    return rows


def run_grep(pattern: str) -> list[str]:
    cmd = f"grep --include='*.adoc' -rn '{pattern}' {GREP_TARGETS}"
    return [l for l in run(cmd).splitlines() if l]


# ---------------------------------------------------------------------------
# Vale output parser
#
# vale --output line produces:
#   path:line:col:severity:message [Rule.Name]
# ---------------------------------------------------------------------------

VALE_RE = re.compile(
    r"^(.+?):(\d+):(\d+):(error|warning|suggestion):(.+?)(?:\s+\[([^\]]+)\])?$"
)


def _parse_vale_line(raw: str) -> dict | None:
    m = VALE_RE.match(raw.strip())
    if not m:
        return None
    return {
        "File": m.group(1),
        "Line": int(m.group(2)),
        "Col": int(m.group(3)),
        "Severity": m.group(4),
        "Message": m.group(5).strip(),
        "Rule": m.group(6) or "",
    }


# ---------------------------------------------------------------------------
# Git sync
# ---------------------------------------------------------------------------

def _git(args: str) -> tuple[int, str, str]:
    result = subprocess.run(
        f"git {args}", shell=True, capture_output=True, text=True
    )
    return result.returncode, result.stdout.strip(), result.stderr.strip()


def sync_with_main() -> bool:
    """Fetch origin and rebase onto origin/main before running checks."""
    code, _, _ = _git("rev-parse --is-inside-work-tree")
    if code != 0:
        print("ERROR: not inside a git repository — cannot sync.")
        return False

    _, branch, _ = _git("rev-parse --abbrev-ref HEAD")
    _, head_sha, _ = _git("rev-parse --short HEAD")
    print(f"Git:         branch '{branch}' at {head_sha}")

    _, status_out, _ = _git("status --porcelain")
    if status_out:
        print(
            "WARNING: uncommitted changes detected — skipping rebase to avoid data loss.\n"
            "         Commit or stash your changes, then re-run.\n"
            "         Checks will run against the current working tree as-is."
        )
        return True

    print("Fetching origin...")
    code, _, err = _git("fetch origin")
    if code != 0:
        print(f"WARNING: git fetch failed ({err}) — running checks against local state.")
        return True

    code, _, _ = _git("rev-parse --verify --quiet origin/main")
    remote_ref = "origin/main" if code == 0 else "origin/master"

    _, behind, _ = _git(f"rev-list --count HEAD..{remote_ref}")
    if behind == "0":
        print(f"Already up to date with {remote_ref}.")
        return True

    print(f"Rebasing onto {remote_ref} ({behind} commit(s) behind)...")
    code, out, err = _git(f"rebase {remote_ref}")
    if code != 0:
        print(
            f"ERROR: rebase failed — resolve conflicts and run again.\n"
            f"  git output: {err or out}"
        )
        _git("rebase --abort")
        return False

    _, new_sha, _ = _git("rev-parse --short HEAD")
    print(f"Rebase complete — now at {new_sha}.")
    return True


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

SEVERITY_COLORS = {
    "error": "FFCCCC",
    "warning": "FFF2CC",
    "suggestion": "CCE5FF",
}
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def write_sheet(ws, headers: list[str], rows: list[dict], col_widths: dict = None):
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([row.get(h, "") for h in headers])
        sev = str(row.get("Severity", "")).lower()
        if sev in SEVERITY_COLORS:
            fill = PatternFill(fill_type="solid", fgColor=SEVERITY_COLORS[sev])
            for cell in ws[ws.max_row]:
                cell.fill = fill

    for col_idx, header in enumerate(headers, 1):
        max_len = max((len(str(r.get(header, ""))) for r in rows), default=0)
        width = col_widths.get(header, min(max_len + 2, 80)) if col_widths else min(max_len + 2, 80)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_summary_sheet(wb, sheets_meta: list[dict]):
    ws = wb.create_sheet("Summary", 0)
    ws.append(["Sheet", "Check", "Count", "Notes"])
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for meta in sheets_meta:
        ws.append([meta["sheet"], meta["check"], meta["count"], meta.get("notes", "")])
    for col_idx, width in enumerate([20, 45, 10, 60], 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="DITA compliance report for MTV / Migration Toolkit for Virtualization"
    )
    parser.add_argument("--vale", default="~/dita.ini", help="Path to vale config")
    parser.add_argument("--out", default="mtv-dita-report.xlsx", help="Output file")
    args = parser.parse_args()

    if not Path("documentation").is_dir():
        sys.exit("ERROR: documentation/ directory not found — run this from the root of the forklift-documentation repo.")

    print("Repo:        MTV / Migration Toolkit for Virtualization (forklift-documentation)")
    print(f"Vale config: {args.vale}")
    print(f"Output:      {args.out}")
    print()

    if not sync_with_main():
        sys.exit(1)
    print()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheets_meta = []

    vale_headers = ["File", "Line", "Col", "Severity", "Message", "Rule"]
    vale_widths = {"File": 55, "Line": 7, "Col": 6, "Severity": 12, "Message": 70, "Rule": 40}

    # ------------------------------------------------------------------
    # Sheet 1 — All DITA errors/warnings
    # ------------------------------------------------------------------
    print("Running: All DITA errors (full vale check)...")
    all_rows = run_vale(args.vale)
    ws1 = wb.create_sheet("All DITA Errors")
    write_sheet(ws1, vale_headers, all_rows, vale_widths)
    sheets_meta.append({
        "sheet": "All DITA Errors",
        "check": "Full vale DITA check",
        "count": len(all_rows),
        "notes": f"{FIND_CMD} | vale --config {args.vale} --output line",
    })
    print(f"  → {len(all_rows)} violations")

    # ------------------------------------------------------------------
    # Sheet 2 — IntrinsicAttribute suggestions
    # ------------------------------------------------------------------
    print("Running: IntrinsicAttribute check...")
    ia_rows = run_vale(args.vale, '.Level=="suggestion" and .Name=="AsciiDocDITA.IntrinsicAttribute"')
    ws2 = wb.create_sheet("IntrinsicAttribute")
    write_sheet(ws2, vale_headers, ia_rows, vale_widths)
    sheets_meta.append({
        "sheet": "IntrinsicAttribute",
        "check": "Intrinsic attribute references (AsciiDocDITA.IntrinsicAttribute)",
        "count": len(ia_rows),
        "notes": "Suggestion-level. Should be 0 for clean DITA migration.",
    })
    print(f"  → {len(ia_rows)} violations")

    # ------------------------------------------------------------------
    # Sheet 3 — AttributeDefinition suggestions
    # ------------------------------------------------------------------
    print("Running: AttributeDefinition check...")
    ad_rows = run_vale(args.vale, '.Level=="suggestion" and .Name=="AsciiDocDITA.AttributeDefinition"')
    ws3 = wb.create_sheet("AttributeDefinition")
    write_sheet(ws3, vale_headers, ad_rows, vale_widths)
    sheets_meta.append({
        "sheet": "AttributeDefinition",
        "check": "Attribute definitions that alter output (AsciiDocDITA.AttributeDefinition)",
        "count": len(ad_rows),
        "notes": "Suggestion-level. Should be 0 for clean DITA migration.",
    })
    print(f"  → {len(ad_rows)} violations")

    # ------------------------------------------------------------------
    # Sheet 4 — TagDirectives suggestions
    # ------------------------------------------------------------------
    print("Running: TagDirectives check...")
    td_rows = run_vale(args.vale, '.Level=="suggestion" and .Name=="AsciiDocDITA.TagDirectives"')
    ws4 = wb.create_sheet("TagDirectives")
    write_sheet(ws4, vale_headers, td_rows, vale_widths)
    sheets_meta.append({
        "sheet": "TagDirectives",
        "check": "Unsupported tag directives (AsciiDocDITA.TagDirectives)",
        "count": len(td_rows),
        "notes": "Suggestion-level. Should be 0 for clean DITA migration.",
    })
    print(f"  → {len(td_rows)} violations")

    # ------------------------------------------------------------------
    # Sheet 5 — Conditional code (ifdef/ifndef/ifeval)
    # ------------------------------------------------------------------
    print("Running: Conditional code checks (grep)...")

    ifeval_lines = run_grep(r"^ifeval::")

    ifdefndef_lines = run_grep(r"^ifn\?def::")
    ifdefndef_non_context = [
        l for l in ifdefndef_lines
        if not re.search(r"\bcontext\b", l.split(":", 2)[2] if len(l.split(":", 2)) > 2 else "")
    ]

    cond_rows = []
    for line in ifeval_lines:
        parts = line.split(":", 2)
        cond_rows.append({
            "File": parts[0] if parts else line,
            "Line": parts[1] if len(parts) > 1 else "",
            "Directive": "ifeval",
            "Content": parts[2] if len(parts) > 2 else "",
            "Note": "ifeval not supported in DITA",
        })
    for line in ifdefndef_non_context:
        parts = line.split(":", 2)
        cond_rows.append({
            "File": parts[0] if parts else line,
            "Line": parts[1] if len(parts) > 1 else "",
            "Directive": "ifdef/ifndef (non-context)",
            "Content": parts[2] if len(parts) > 2 else "",
            "Note": "Review — context-based conditionals are supported; others may not be",
        })

    cond_headers = ["File", "Line", "Directive", "Content", "Note"]
    cond_widths = {"File": 55, "Line": 7, "Directive": 25, "Content": 60, "Note": 55}
    ws5 = wb.create_sheet("Conditional Code")
    write_sheet(ws5, cond_headers, cond_rows, cond_widths)
    sheets_meta.append({
        "sheet": "Conditional Code",
        "check": "ifeval directives + non-context ifdef/ifndef",
        "count": len(cond_rows),
        "notes": f"ifeval: {len(ifeval_lines)}, non-context ifdef/ifndef: {len(ifdefndef_non_context)}",
    })
    print(f"  → {len(cond_rows)} conditional directives ({len(ifeval_lines)} ifeval, {len(ifdefndef_non_context)} non-context ifdef/ifndef)")

    # ------------------------------------------------------------------
    # Sheet 6 — Conditional attribute breakdown
    # ------------------------------------------------------------------
    print("Running: Conditional attribute breakdown...")

    attr_counts: dict[str, int] = {}
    for line in ifdefndef_lines:
        parts = line.split(":", 2)
        content = parts[2] if len(parts) > 2 else ""
        attr_str = re.sub(r"^ifn?def::", "", content.strip()).split("[")[0]
        for attr in re.split(r"[,+]", attr_str):
            attr = attr.strip()
            if attr and not re.search(r"\bcontext\b", attr):
                attr_counts[attr] = attr_counts.get(attr, 0) + 1

    attr_rows = [
        {"Count": count, "Attribute": attr}
        for attr, count in sorted(attr_counts.items(), key=lambda x: -x[1])
    ]
    attr_headers = ["Count", "Attribute"]
    attr_widths = {"Count": 10, "Attribute": 40}
    ws6 = wb.create_sheet("Conditional Attributes")
    write_sheet(ws6, attr_headers, attr_rows, attr_widths)
    sheets_meta.append({
        "sheet": "Conditional Attributes",
        "check": "Attribute names used in non-context ifdef/ifndef conditions",
        "count": len(attr_rows),
        "notes": "High counts = widespread conditional; review each for DITA compatibility",
    })
    print(f"  → {len(attr_rows)} distinct non-context attributes")

    # ------------------------------------------------------------------
    # Sheet 0 — Summary (inserted at front)
    # ------------------------------------------------------------------
    write_summary_sheet(wb, sheets_meta)

    wb.save(args.out)
    print(f"\nReport written to: {args.out}")
    print(f"Total issues across all checks: {sum(m['count'] for m in sheets_meta)}")


if __name__ == "__main__":
    main()
